"""
双表头/多数据块复杂布局手动格式脚本 — 模板

适用场景
--------
beautify 无法处理的复杂布局：
- 大标题(Row2) + 列头(Row3) + 数据(Row4) + 第二列头(Row6) + 数据(Row7+)
- 横向两段独立表格（第一段从A列、第二段从E列）
- 合并单元格 + 多级表头

使用方式
--------
1. 复制此文件到目标目录
2. 修改 path 为实际文件路径
3. 根据实际表格结构调整 section_rows 和 header_rows
4. 修改 theme_name 为目标主题
5. python dual-header-format.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from shutil import copy2

# ═════════════════════════════════════════════════════
# 配置区 — 按实际表格修改
# ═════════════════════════════════════════════════════

path = "目标文件.xlsx"
theme_name = "default"  # 可选: default, deep-navy, jade, slate, burgundy, coral,
                        #       sakura, warm-sun, lavender, matcha, peach, misty

# 表格区块定义：[(区块起始行, 区块结束行, 表头行), ...]
# 起始行/结束行 包含表头和数据
sections = [
    # (start_row, end_row, header_row)
    (3, 10, 3),    # 第一段：Row3=列头, Row4-10=数据
    (6, 17, 6),    # 第二段：Row6=列头, Row7-17=数据
]

# 数据列范围
start_col = 2   # B列
end_col = 12    # L列

# 冻结行（可选）
freeze_cell = "B4"

# ═════════════════════════════════════════════════════
# 主题色定义（同 make_excel.py）
# ═════════════════════════════════════════════════════

THEMES = {
    'default': {'header':'4472C4','data':'2C5F8A','sum':'D9E2F3'},
    'burgundy': {'header':'843C0C','data':'5B2C0A','sum':'FCE4D6'},
    'sakura': {'header':'D94F70','data':'8B2A3E','sum':'FDE8EE'},
    'warm-sun': {'header':'E8843A','data':'7B3F00','sum':'FEF0E6'},
    'lavender': {'header':'8B7EC8','data':'4A3F7A','sum':'EDEAF8'},
    'matcha': {'header':'7BA23F','data':'3D5A1E','sum':'F0F5E6'},
    'peach': {'header':'E8897C','data':'8B3A2E','sum':'FEF0ED'},
    'misty': {'header':'6B7FB5','data':'3D4A6B','sum':'EDF0F8'},
    'deep-navy': {'header':'1F4E79','data':'163A5E','sum':'D6E4F0'},
    'jade': {'header':'375623','data':'2A4219','sum':'E2EFDA'},
    'slate': {'header':'404040','data':'2D2D2D','sum':'D9D9D9'},
    'coral': {'header':'D84B4B','data':'8B2E2E','sum':'FDE8E8'},
}

# ═════════════════════════════════════════════════════
# 格式常量
# ═════════════════════════════════════════════════════

THICK = Side(style='medium', color='000000')
DASHED = Side(style='dashed', color='808080')
NO = Side(style=None)

TOP_BORDER = Border(top=THICK, bottom=DASHED, left=NO, right=NO)
MID_BORDER = Border(top=DASHED, bottom=DASHED, left=NO, right=NO)
BOTTOM_BORDER = Border(top=DASHED, bottom=THICK, left=NO, right=NO)

ROW_HEIGHT = 18
HDR_ALIGN = Alignment(horizontal='right', vertical='center')

# ═════════════════════════════════════════════════════
# 执行
# ═════════════════════════════════════════════════════

t = THEMES.get(theme_name, THEMES['default'])
hd_fill = PatternFill(start_color=t['header'], end_color=t['header'], fill_type='solid')
hd_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
data_font = Font(name='Arial', size=11, color=t['data'])

# 备份
copy2(path, path.replace('.xlsx', f'_backup_manual.xlsx'))

wb = openpyxl.load_workbook(path)
ws = wb.active

for sec_start, sec_end, hdr_row in sections:
    ncols = end_col - start_col + 1
    
    # 表头
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=hdr_row, column=c)
        cell.fill = hd_fill
        cell.font = hd_font
        cell.alignment = HDR_ALIGN
    
    # 数据行
    for r in range(sec_start, sec_end + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # 边框
    data_rows = list(range(hdr_row, sec_end + 1))
    for ri, r in enumerate(data_rows):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if ri == 0:
                cell.border = TOP_BORDER
            elif ri == len(data_rows) - 1:
                cell.border = BOTTOM_BORDER
            else:
                cell.border = MID_BORDER

ws.sheet_view.showGridLines = False
if freeze_cell:
    ws.freeze_panes = freeze_cell

wb.save(path)
print(f"Done. Formatted {len(sections)} section(s) with '{theme_name}' theme.")
