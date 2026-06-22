#!/usr/bin/env python3
"""
make_excel — 从 DataFrame 生成摩根士丹利标准格式的 Excel 报表。
beautify   — 对已有 Excel 文件只改格式不改数据（保留公式和值）。

设计哲学
--------
这套脚本把《为什么精英都是Excel控》这本书的格式规范实现了。
核心原则：不需要用户学100条Excel技巧，调8个设置就够了。

两个入口：
- make_excel(df, path)    : 从零生成。写数据→上格式→估列宽→一次保存。
- beautify(path, output)  : 美化已有文件。不改任何单元格的值，只改格式。
  保留公式，公式计算结果用黑色字体，手动输入的值用蓝色字体（摩根系标准）。

为什么不用 xlwings？
xlwings 保存时会丢掉 openpyxl 的边框/颜色/数字格式。
之前为了 autofit 做了三步修补循环（写→保存→xlwings打开→读列宽→写回→再保存），
这是架构级别的脆弱性。改为字符估算法后，一次保存完事。

为什么不做 autofit？
openpyxl 没有 autofit 方法。xlwings 有但会吞格式。
字符估算法（中文字符计2，拉丁/数字计1）足够准确，且不依赖外部进程。

依赖: pandas, openpyxl>=3.0（纯 openpyxl，无 xlwings）
设计: 写数据 → 上格式 → 估列宽 → 一次保存，线性流程不回头。
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path
from shutil import copy2
from typing import Union, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════
# 格式常量 — 摩根士丹利标准
#
# 全书核心就8条规则，这里全部落成代码。
# ═══════════════════════════════════════════════════════════════════════════

# ── 框线（规则5：上下粗线 + 中间虚线，不要竖线）──
# 摩根系认为：竖线会让表看起来像牢笼。边框只做横向分割。
THICK = Side(style='medium', color='000000')   # 上下边界用粗线
DASHED = Side(style='dashed', color='808080')  # 中间分隔用虚线
NO = Side(style=None)                           # 左右/竖边不留线

TOP_BORDER = Border(top=THICK, bottom=DASHED, left=NO, right=NO)
MID_BORDER = Border(top=DASHED, bottom=DASHED, left=NO, right=NO)
BOTTOM_BORDER = Border(top=DASHED, bottom=THICK, left=NO, right=NO)

# ── 表头（规则6：水蓝底 #4472C4 + 白字粗体）──
HDR_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HDR_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
HDR_ALIGN = Alignment(horizontal='right', vertical='center')  # 表头对齐方式跟随数据列

# ── 数据字体颜色 ──
# 摩根系规范：手动输入的数字用蓝色（代表可调的价值动因），公式计算结果用黑色。
# make_excel 模式：数据来自 pandas，都是值（无公式），所以数字列全部蓝色。
# beautify 模式：双重检测公式（str.startswith('='）+ cell.data_type == 'f'）。
DATA_FONT_BLUE = Font(name='Arial', size=11, color='0000FF')
DATA_FONT_BLACK = Font(name='Arial', size=11, color='000000')

# ── 合计行（自动检测 + 特殊样式）──
# 检测逻辑：最后一行首列（A或B列）包含合计关键词 → 整行加浅蓝底粗体。
# 关键词匹配是包含关系（str in str），不是精准匹配。
SUM_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
SUM_FONT = Font(name='Arial', size=11, bold=True)
SUM_KEYWORDS = ('合计', '小计', '总计', 'sum', 'total', '小計', '合計')

# ── 数据对齐（规则4：文字左对齐，数字右对齐）──
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')

# ── 数字格式（规则3：千分位符）──
# 根据列类型自动匹配数字格式。
# fmt_override 参数可覆盖默认格式，如 {'money': '#,##0.0', 'pct': '0.0%'}。
# 注意：money 默认两位小数（#,##0.00），整数用 #,##0。
FMT = {
    'money': '#,##0.00',
    'number': '#,##0',
    'pct': '0.00%',
    'date': 'yyyy/mm/dd',
    'text': '@',
    'unknown': '@',
}

# ── 布局（规则1：行高18，规则7：A列留空）──
ROW_HEIGHT = 18            # 默认13.5太挤，中文笔画密，18 才有呼吸感
A_COL_WIDTH = 2             # A列留空，左侧视觉缓冲
MIN_COL_WIDTH = 10          # 列宽下限（太窄文字折行）
MAX_COL_WIDTH = 50          # 列宽上限（太宽影响一览性）


# ═══════════════════════════════════════════════════════════════════════════
# 色系主题
#
# 6 套配色，满足不同场景和审美偏好。
# 主题只改颜色不改布局——框线规则、对齐方式、数字格式不受主题影响。
# 新增主题：在 THEMES 字典中加一项即可，无需改其他代码。
# 注意：theme 不存在时会抛 ValueError，不会静默回退到 default。
# ═══════════════════════════════════════════════════════════════════════════

THEMES = {
    # ── 经典商务系（冷色/中性） ──
    'default': {
        'label': '水蓝',
        'header_fill': '4472C4',
        'header_font_color': 'FFFFFF',
        'sum_fill': 'D9E2F3',
        'sum_font_bold': True,
        'sum_font_color': '000000',
        'data_font': '2C5F8A',      # 深蓝，匹配表头
        'formula_font': '000000',
        'desc': '经典水蓝（原摩根系默认）',
    },
    'deep-navy': {
        'label': '深海蓝',
        'header_fill': '1F4E79',
        'header_font_color': 'FFFFFF',
        'sum_fill': 'D6E4F0',
        'sum_font_bold': True,
        'sum_font_color': '000000',
        'data_font': '163A5E',
        'formula_font': '000000',
        'desc': '深海蓝表头，更沉稳专业',
    },
    'jade': {
        'label': '墨玉绿',
        'header_fill': '375623',
        'header_font_color': 'FFFFFF',
        'sum_fill': 'E2EFDA',
        'sum_font_bold': True,
        'sum_font_color': '000000',
        'data_font': '2A4219',
        'formula_font': '000000',
        'desc': '墨绿表头，清爽自然',
    },
    'slate': {
        'label': '陨石灰蓝',
        'header_fill': '404040',
        'header_font_color': 'FFFFFF',
        'sum_fill': 'D9D9D9',
        'sum_font_bold': True,
        'sum_font_color': '000000',
        'data_font': '2D2D2D',
        'formula_font': '000000',
        'desc': '深灰表头，现代极简',
    },

    # ── 暖色高级感系 ──
    'burgundy': {
        'label': '勃艮第红',
        'header_fill': '843C0C',
        'header_font_color': 'FFFFFF',
        'sum_fill': 'FCE4D6',
        'sum_font_bold': True,
        'sum_font_color': '000000',
        'data_font': '5B2C0A',      # 深褐，匹配酒红
        'formula_font': '000000',
        'desc': '酒红表头，暖色调高级感',
    },
    'coral': {
        'label': '珊瑚橙',
        'header_fill': 'D84B4B',
        'header_font_color': 'FFFFFF',
        'sum_fill': 'FDE8E8',
        'sum_font_bold': True,
        'sum_font_color': '000000',
        'data_font': '8B2E2E',
        'formula_font': '000000',
        'desc': '珊瑚红表头，活泼明亮',
    },

    # ── 青春活力系（新增 6 套） ──
    # 整体调性：暖色、饱和度不过高、视觉协调不扎眼
    'sakura': {
        'label': '樱花粉',
        'header_fill': 'D94F70',
        'header_font_color': 'FFFFFF',
        'sum_fill': 'FDE8EE',
        'sum_font_bold': True,
        'sum_font_color': '000000',
        'data_font': '8B2A3E',      # 深浆果色
        'formula_font': '000000',
        'desc': '粉红表头，温柔不失活力',
    },
    'warm-sun': {
        'label': '暖阳橙',
        'header_fill': 'E8843A',
        'header_font_color': 'FFFFFF',
        'sum_fill': 'FEF0E6',
        'sum_font_bold': True,
        'sum_font_color': '000000',
        'data_font': '7B3F00',      # 深琥珀
        'formula_font': '000000',
        'desc': '暖橙表头，阳光有能量',
    },
    'lavender': {
        'label': '薰衣草紫',
        'header_fill': '8B7EC8',
        'header_font_color': 'FFFFFF',
        'sum_fill': 'EDEAF8',
        'sum_font_bold': True,
        'sum_font_color': '000000',
        'data_font': '4A3F7A',      # 深紫
        'formula_font': '000000',
        'desc': '淡紫表头，知性温婉',
    },
    'matcha': {
        'label': '抹茶绿',
        'header_fill': '7BA23F',
        'header_font_color': 'FFFFFF',
        'sum_fill': 'F0F5E6',
        'sum_font_bold': True,
        'sum_font_color': '000000',
        'data_font': '3D5A1E',      # 深茶绿
        'formula_font': '000000',
        'desc': '抹茶绿表头，清新自然',
    },
    'peach': {
        'label': '蜜桃',
        'header_fill': 'E8897C',
        'header_font_color': 'FFFFFF',
        'sum_fill': 'FEF0ED',
        'sum_font_bold': True,
        'sum_font_color': '000000',
        'data_font': '8B3A2E',      # 深桃红褐
        'formula_font': '000000',
        'desc': '蜜桃粉表头，甜美不腻',
    },
    'misty': {
        'label': '雾蓝紫',
        'header_fill': '6B7FB5',
        'header_font_color': 'FFFFFF',
        'sum_fill': 'EDF0F8',
        'sum_font_bold': True,
        'sum_font_color': '000000',
        'data_font': '3D4A6B',      # 深蓝紫
        'formula_font': '000000',
        'desc': '雾蓝紫表头，现代感冷色调',
    },
}

AVAILABLE_THEMES = ', '.join(sorted(THEMES.keys()))


# ═══════════════════════════════════════════════════════════════════════════
# 主题校验 + 构建
# ═══════════════════════════════════════════════════════════════════════════

def _validate_theme(theme_name: str):
    """
    校验主题名，无效则抛 ValueError。
    这是防止静默回退的保险——调用方以为用了 coral，实际是 default，这是已知 bug。
    两层校验：_build_theme_styles 内部也调一次，确保不会漏。
    """
    if theme_name not in THEMES:
        raise ValueError(
            f"未知主题 '{theme_name}'。可用主题: {sorted(THEMES.keys())}"
        )


def _build_theme_styles(theme_name: str):
    """
    根据主题名构建 PatternFill / Font 对象。
    二次校验 theme_name（即使调用方已经校验过）。
    """
    _validate_theme(theme_name)
    t = THEMES.get(theme_name, THEMES['default'])  # get 是兜底，validate 已确保命中
    return {
        'header_fill': PatternFill(start_color=t['header_fill'], end_color=t['header_fill'], fill_type='solid'),
        'header_font': Font(name='Arial', size=11, bold=True, color=t['header_font_color']),
        'sum_fill': PatternFill(start_color=t['sum_fill'], end_color=t['sum_fill'], fill_type='solid'),
        'sum_font': Font(name='Arial', size=11, bold=t['sum_font_bold'], color=t['sum_font_color']),
        'data_font': Font(name='Arial', size=11, color=t['data_font']),
        'formula_font': Font(name='Arial', size=11, color=t['formula_font']),
    }


# ═══════════════════════════════════════════════════════════════════════════
# 列类型推断
#
# make_excel 模式用 _infer_column_type（依赖 pandas Series）
# beautify  模式用 _infer_col_type_from_ws（独立于 pandas，从单元格采样）
# 两套逻辑的关键词匹配部分保持一致，值特征部分因数据源不同各有侧重。
#
# 优先级：pct > date > text > money
# 为什么 pct 最优先？因为"毛利率"这个词同时匹配 pct（率）和 money（毛利），
# 如果 money 先匹配就会被错误归类。用 `率$` 限定词尾 + `毛利(?!率)` 排除误匹配。
# ═══════════════════════════════════════════════════════════════════════════

def _infer_column_type(name: str, series: pd.Series) -> str:
    """
    根据列名关键词 + 值特征推断列类型。
    用于 make_excel 模式（有 pandas Series 可用）。

    参数
    ----
    name : str — 列名
    series : pd.Series — 该列数据

    返回
    ----
    str : 类型名（money/number/pct/date/text/unknown）

    注意
    ----
    类型推断不是100%准确的。如果 DataFrame 的列名含糊（如"数据1""字段A"），
    关键词匹配会落空，完全依赖值特征。此时文本列可能被误判为 money。
    调用方可以事后用 fmt_override 修正。
    """
    n = name.lower()

    # ── 关键词匹配（优先级：pct > date > text > money）──
    # "率"优先检测，避免"毛利率"被 money 关键词误匹配
    if re.search(r'(率$|占比|百分比|增长率|rate$|ratio$)', n):
        return 'pct'
    if re.search(r'(日期|时间|年月|期间|年份|月份|date|time|year|month)', n):
        return 'date'
    if re.search(r'(编号|id|单号|编码|电话|手机|备注|说明|名称|地址|描述|号码|订单号|负责人|姓名|联系人|部门|岗位|phone|email|code|desc|address)', n):
        return 'text'
    # 注意：毛利(?!率) 匹配"毛利"但不匹配"毛利率"（已归为 pct）
    if re.search(r'(金额|价格|收入|成本|费用|毛利额|净利|合计|总额|售价|单价|预算|支出|毛利(?!率)|amount|price|cost|revenue|total|budget|expense)', n):
        return 'money'

    # ── 值特征 ──
    clean = series.dropna()
    if len(clean) == 0:
        return 'unknown'

    try:
        if pd.api.types.is_datetime64_any_dtype(clean):
            return 'date'

        if pd.api.types.is_numeric_dtype(clean):
            # 百分比只靠列名关键词推断，不依赖值范围。
            # 0~1 的浮点可能是万元/千元单位，不是百分数。
            if pd.api.types.is_float_dtype(clean):
                return 'money'
            return 'number'

        # 长数字字符串（如身份证号、订单号）→ text，防止被当作 number
        str_sample = clean.astype(str).head(20)
        if str_sample.str.match(r'^\d{12,}$').any():
            return 'text'
    except Exception:
        pass

    return 'unknown'


# ═══════════════════════════════════════════════════════════════════════════
# 列宽估算
#
# openpyxl 没有 autofit。xlwings 有但会吞格式。
# 改用字符估算法：中文字符计2，拉丁/数字计1，取最宽值 + 2 padding。
# 不算精确，但够用。如果用户对列宽有严格要求，可以在 beautify 之后手动调。
# ═══════════════════════════════════════════════════════════════════════════

def _cell_display_text(cell) -> str:
    """
    安全提取单元格的展示文本。
    公式对象（ArrayFormula/TableFormula）不能直接用 str() 转，会得到对象内存地址。
    数值/字符串/日期/None 都能正确处理。
    """
    v = cell.value if hasattr(cell, 'value') else cell
    if v is None:
        return ''
    if isinstance(v, str):
        return v
    # openpyxl formula objects (ArrayFormula, SharedFormula, etc.)
    if hasattr(v, 'value'):
        fv = v.value
        if isinstance(fv, str):
            return fv
    # DataTableFormula — Excel 模拟运算表公式，无公式字符串可提取
    # 这类单元格在 Excel 中显示的是计算结果，不是公式文本
    # 返回空字符串让列宽估算走其他单元格或走默认
    return ''


def _char_width(text: str) -> int:
    """估算字符串宽度：中文字符计2，拉丁/数字计1。"""
    return sum(2 if ord(ch) > 127 else 1 for ch in str(text))


def _estimate_column_width(header: str, series: pd.Series) -> int:
    """
    估算列宽：取表头和数据中最大字符宽度 + padding。
    用于 make_excel 模式。
    """
    max_w = _char_width(header)
    for v in series.dropna():
        try:
            w = _char_width(str(v))
            if w > max_w:
                max_w = w
        except Exception:
            pass
    return max(MIN_COL_WIDTH, min(max_w + 2, MAX_COL_WIDTH))


def _estimate_col_width_from_cells(ws, col, header_row, data_end):
    """
    从工作表的表头和单元格值估算列宽（安全处理公式对象）。
    用于 beautify 模式（不依赖 pandas，直接从工作表取值）。
    """
    max_w = _char_width(_cell_display_text(ws.cell(row=header_row, column=col)))
    for r in range(header_row + 1, data_end + 1):
        txt = _cell_display_text(ws.cell(row=r, column=col))
        if txt:
            try:
                w = _char_width(txt)
                if w > max_w:
                    max_w = w
            except Exception:
                pass
    return max(MIN_COL_WIDTH, min(max_w + 2, MAX_COL_WIDTH))


# ═══════════════════════════════════════════════════════════════════════════
# 格式引擎 — make_excel 模式
#
# 写数据 + 应用格式，线性流程一次保存。
# 数据从 B2 开始写，A列留空（规则7）。
# ═══════════════════════════════════════════════════════════════════════════

def _apply_styles(ws, df: pd.DataFrame, theme: str = 'default', fmt_override: Optional[dict] = None,
                   freeze_rows: Optional[int] = None):
    """
    对已写入数据的工作表应用摩根系格式。

    流程：
    1. 行高 + A列宽
    2. 列宽估算 + 类型推断（一次扫描）
    3. 表头：水蓝底白字右对齐
    4. 数据行：统一 Arial 11 + 对齐 + 数字格式 + 字体颜色
    5. 合计行检测（最后一行首列含合计关键词→特殊样式）
    6. 边框（上下粗中间虚线无竖线）
    7. 右侧空白列（宽3）
    8. 隐藏网格线 + 冻结表头

    参数
    ----
    fmt_override : dict, optional
        自定义数字格式，如 {'money': '#,##0.0', 'pct': '0.0%'}。
    freeze_rows : int, optional
        要冻结的行数。默认 None 自动推断：make_excel 模式表头在第1行，冻结1行（A2）。
    """
    s = _build_theme_styles(theme)
    # 合并自定义数字格式
    effective_fmt = dict(FMT)
    if fmt_override:
        effective_fmt.update(fmt_override)
    nrows = 1 + len(df)   # 表头 + 数据行
    ncols = len(df.columns)
    start_col = 2         # B 列（A列留空）

    # 1. 行高 + A列宽
    for r in range(1, nrows + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT
    ws.column_dimensions['A'].width = A_COL_WIDTH

    # 2. 列宽估算 + 类型推断（一次扫描）
    col_types = {}
    for ci in range(ncols):
        col_name = df.columns[ci]
        col_letter = get_column_letter(ci + start_col)
        col_type = _infer_column_type(str(col_name), df.iloc[:, ci])
        col_types[ci] = col_type

        est = _estimate_column_width(str(col_name), df.iloc[:, ci])
        # 数值/百分比列如果含长公式，固定列宽15（make_excel 模式数据来自 pandas，一般无公式，但兜底处理）
        if col_type in ('money', 'number', 'pct'):
            for v in df.iloc[:, ci].dropna():
                sv = str(v)
                if sv.startswith('=') and len(sv) > 15:
                    est = 15
                    break
        ws.column_dimensions[col_letter].width = est

    # 3. 表头（第1行）
    for ci in range(ncols):
        cell = ws.cell(row=1, column=ci + start_col)
        cell.font = s['header_font']
        cell.fill = s['header_fill']
        cell.alignment = HDR_ALIGN

    # 4. 数据行
    # 统一 Arial 11（规则2），数字右对齐（规则4），数字格式按类型匹配（规则3）
    for ci in range(ncols):
        ct = col_types.get(ci, 'unknown')
        is_num = ct in ('money', 'number', 'pct')
        align = RIGHT_ALIGN if is_num else LEFT_ALIGN
        nf = effective_fmt.get(ct)

        for ri in range(2, nrows + 1):
            cell = ws.cell(row=ri, column=ci + start_col)
            # 所有单元格统一 Arial 11
            cell.font = s['data_font'] if is_num else Font(name='Arial', size=11)
            cell.alignment = align
            if nf:
                cell.number_format = nf

    # 5. 合计行检测与特殊格式
    # 检测最后一行的首列（第一数据列）是否含合计关键词
    last_val = ws.cell(row=nrows, column=start_col).value
    is_summary = (
        nrows > 1
        and last_val is not None
        and any(kw in str(last_val) for kw in SUM_KEYWORDS)
    )
    if is_summary:
        for ci in range(ncols):
            cell = ws.cell(row=nrows, column=ci + start_col)
            cell.font = s['sum_font']
            cell.fill = s['sum_fill']

    # 6. 边框（规则5：上下粗中间虚线无竖线）
    for ci in range(ncols):
        col = ci + start_col
        for ri in range(1, nrows + 1):
            cell = ws.cell(row=ri, column=col)
            if ri == 1:
                cell.border = TOP_BORDER
            elif ri == nrows:
                cell.border = BOTTOM_BORDER
            else:
                cell.border = MID_BORDER

    # 6b. 右侧空白列（宽3），防止数据抵到表格边缘
    last_data_col = start_col + ncols - 1
    right_col = get_column_letter(last_data_col + 1)
    ws.column_dimensions[right_col].width = 3

    # 7. 全局设置：隐藏网格线 + 冻结表头
    # 4 种冻结策略（优先级由高到低）：
    #   1. freeze_rows 参数显式指定 → 用指定值
    #   2. beautify 模式检测到的 header_row → 冻结到表头行
    #   3. B2起始布局（行1留空，行2表头）→ 冻结3，即冻结行1-2
    #   4. 默认：表头在第1行 → 冻结2，即冻结行1
    ws.sheet_view.showGridLines = False
    if freeze_rows is not None:
        freeze_cell = f'A{freeze_rows + 1}'
    elif ws.cell(row=1, column=2).value and not ws.cell(row=1, column=1).value:
        # B2起始：行1只有B列有值（表头），A列空 → 行1本身是数据表头，冻结A2
        freeze_cell = 'A2'
    else:
        freeze_cell = 'A2'
    ws.freeze_panes = freeze_cell


# ═══════════════════════════════════════════════════════════════════════════
# 写数据
#
# 数据从 B2 开始写，A列留空（规则7）。
# 第1行 = 表头，第2行起 = 数据。
# ═══════════════════════════════════════════════════════════════════════════

def _write_data(ws, df: pd.DataFrame):
    """从 B2 开始写入（第1行 = 表头，A列留空）。"""
    for ci, col_name in enumerate(df.columns):
        ws.cell(row=1, column=ci + 2, value=col_name)
    for ri, (_, row) in enumerate(df.iterrows()):
        for ci, value in enumerate(row):
            ws.cell(row=ri + 2, column=ci + 2, value=value)


# ═══════════════════════════════════════════════════════════════════════════
# 主入口 — make_excel
# ═══════════════════════════════════════════════════════════════════════════

def make_excel(data, output_path: str, sheet_name: str = 'Sheet1', theme: str = 'default',
               fmt_override: Optional[dict] = None,
               freeze_rows: Optional[int] = None) -> str:
    """
    生成摩根系标准格式 Excel。

    参数
    ----
    data : DataFrame | list[tuple[str, DataFrame]]
        单 sheet 传 df，多 sheet 传 [(名称, df), ...]。
    output_path : str
        输出文件路径。
    sheet_name : str
        单 sheet 模式的工作表名称。
    theme : str
        色系主题名。可用: """ + AVAILABLE_THEMES + """。默认 'default'。
    fmt_override : dict, optional
        自定义数字格式，如 {'money': '#,##0.0', 'pct': '0.0%'}。
        合并到默认 FMT 之上，同名覆盖。
    freeze_rows : int, optional
        要冻结的行数。默认 None 自动推断。
        make_excel 模式：表头在第1行→冻1行。传 0 则不冻结。

    返回
    ----
    str : 输出文件的绝对路径。

    注意
    ----
    make_excel 模式下所有数字都是蓝色（假定为手动输入）。
    如果源数据包含公式计算结果，用 beautify 替代。
    """
    _validate_theme(theme)

    # 统一为多 sheet 格式
    if isinstance(data, pd.DataFrame):
        sheets = [(sheet_name, data)]
    else:
        sheets = list(data)

    wb = Workbook()
    for name in list(wb.sheetnames):
        del wb[name]

    for idx, (name, df) in enumerate(sheets):
        if df is None or (hasattr(df, 'empty') and df.empty):
            continue
        # 工作表名称不能超过31字符，不能含特殊字符
        safe_name = re.sub(r'[\\/*?:\[\]]', '_', str(name))[:31]
        ws = wb.create_sheet(title=safe_name, index=idx)
        _write_data(ws, df)
        _apply_styles(ws, df, theme=theme, fmt_override=fmt_override, freeze_rows=freeze_rows)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out.absolute())


# ═══════════════════════════════════════════════════════════════════════════
# beautify — 只改格式不改数据
#
# beautify 的核心差异：
# 1. 不写数据：所有 cell.value 不动，公式原样残留
# 2. 数据检测：_detect_data_range() 扫描工作表找实际数据范围
# 3. 类型推断：从单元格采样，不依赖 pandas Series
# 4. 公式判断：双重检测 str.startswith('='）+ cell.data_type == 'f'
# 5. 备份机制：原地覆盖前自动备份原文件
# ═══════════════════════════════════════════════════════════════════════════

def _detect_data_range(ws):
    """
    检测工作表的实际数据范围。

    返回 (header_row, data_start, data_end, col_end) 或 None（空表）。

    检测规则：
    - 从第1行开始找第一个非空行作为表头行
    - 从表头行之后找最后一个非空行作为数据结束
    - 跳过列A（索引列通常为空）

    已知限制：
    - 如果工作表前 N 行有零散的非表头数据（如备注行、空行混杂注释），
      可能误判 header_row。beautify 时如果遇到 A 列全空且其他列有数据的表格，
      使用前确认首行确实是表头。
    """
    max_col = 0
    header_row = None
    last_data_row = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1):
        r = row[0].row
        has_data = any(cell.value is not None for cell in row)
        if has_data and header_row is None:
            header_row = r
        if has_data:
            last_data_row = r
        for cell in row:
            if cell.value is not None and cell.column > max_col:
                max_col = cell.column

    if header_row is None:
        return None  # 空表

    data_start = header_row + 1
    data_end = max(last_data_row, data_start - 1)
    start_col = 2  # 从 B 列开始（A列留空规范）
    end_col = max(start_col, max_col)
    return header_row, data_start, data_end, start_col, end_col


def _sample_ws_column(ws, col, start_row, end_row, max_samples=100):
    """
    从工作表列中采样非空值，用于类型推断。
    上限100条，够判断趋势就行，不扫全表（大文件性能考虑）。
    """
    samples = []
    for r in range(start_row, end_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None:
            samples.append(v)
            if len(samples) >= max_samples:
                break
    return samples


def _infer_col_type_from_ws(header_value, samples):
    """
    从工作表列推断类型（不依赖 pandas Series）。
    用于 beautify 模式。关键词匹配逻辑与 _infer_column_type 保持一致。

    参数
    ----
    header_value : str — 表头单元格的值
    samples : list — 该列数据采样

    返回
    ----
    str : 类型名（money/number/pct/date/text/unknown）
    """
    name = str(header_value) if header_value is not None else ''
    n = name.lower()

    # 关键词匹配（优先级：pct > date > text > money）
    if re.search(r'(率$|占比|百分比|增长率|rate$|ratio$)', n):
        return 'pct'
    if re.search(r'(日期|时间|年月|期间|年份|月份|date|time|year|month)', n):
        return 'date'
    if re.search(r'(编号|id|单号|编码|电话|手机|备注|说明|名称|地址|描述|号码|订单号|负责人|姓名|联系人|部门|岗位|phone|email|code|desc|address)', n):
        return 'text'
    if re.search(r'(金额|价格|收入|成本|费用|毛利额|净利|合计|总额|售价|单价|预算|支出|毛利(?!率)|amount|price|cost|revenue|total|budget|expense)', n):
        return 'money'

    if not samples:
        return 'unknown'

    # 值采样分析
    numeric_samples = [v for v in samples if isinstance(v, (int, float))]
    str_samples = [str(v) for v in samples if not isinstance(v, (int, float))]

    # 检查日期类型（datetime/date 对象过半则判为 date）
    from datetime import datetime, date
    date_count = sum(1 for v in samples if isinstance(v, (datetime, date)))
    if date_count >= len(samples) * 0.5:
        return 'date'

    if numeric_samples:
        float_samples = [v for v in numeric_samples if isinstance(v, float)]
        if float_samples:
            # 百分比只靠列名关键词推断，值在 0~1 不自动视为百分数
            return 'money'
        return 'number'

    # 长数字字符串 → text（如身份证号、订单号）
    if any(re.match(r'^\d{12,}$', s) for s in str_samples):
        return 'text'

    return 'unknown'


def _beautify_worksheet(ws, col_types_override: Optional[dict] = None, theme: str = 'default',
                        fmt_override: Optional[dict] = None,
                        freeze_rows: Optional[int] = None):
    """
    对单个工作表应用摩根系格式（不改变任何单元格的值）。

    参数
    ----
    col_types_override : dict, optional
        强制指定列的类型，如 {'订单号': 'text', '金额': 'money'}。
        列名匹配，覆盖自动推断结果。
    theme : str
        色系主题名。可用: """ + AVAILABLE_THEMES + """。默认 'default'。
    fmt_override : dict, optional
        自定义数字格式，如 {'money': '#,##0.0', 'pct': '0.0%'}。
    freeze_rows : int, optional
        要冻结的行数。默认 None 自动推断：根据检测到的 header_row 冻结。
        传 0 则不冻结。

    与 make_excel 的差异：
    - 数据范围动态检测（不是固定从 B2 开始）
    - 类型推断从单元格采样（不依赖 pandas Series）
    - 公式双重检测（startswith('='）+ data_type == 'f'）
    - A 列自适应宽度（有数据时补齐摩根系格式）
    """
    s = _build_theme_styles(theme)
    effective_fmt = dict(FMT)
    if fmt_override:
        effective_fmt.update(fmt_override)
    dr = _detect_data_range(ws)
    if dr is None:
        return  # 空表跳过
    header_row, data_start, data_end, start_col, end_col = dr

    # 1. 行高 + A列宽
    # A 列如果有数据，自适应宽度（下限10，上限50）；否则保持 2（留空）
    for r in range(header_row, data_end + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT
    a_header_val = ws.cell(row=header_row, column=1).value
    if a_header_val is not None:
        a_max = _char_width(str(a_header_val))
        for r in range(data_start, data_end + 1):
            v = ws.cell(row=r, column=1).value
            if v is not None:
                try:
                    w = _char_width(str(v))
                    if w > a_max:
                        a_max = w
                except Exception:
                    pass
        ws.column_dimensions['A'].width = max(MIN_COL_WIDTH, min(a_max + 2, MAX_COL_WIDTH))
    else:
        ws.column_dimensions['A'].width = A_COL_WIDTH

    # 2. 列宽估算 + 类型推断（支持手动覆盖）
    col_types = {}
    col_headers = {}  # 列号 → 列名，用于 override 匹配
    for c in range(start_col, end_col + 1):
        col_letter = get_column_letter(c)
        header_val = ws.cell(row=header_row, column=c).value
        col_headers[c] = str(header_val) if header_val is not None else ''

        # 手动覆盖优先——当自动推断不准时，调用方可以精确指定
        if col_types_override and col_headers[c] in col_types_override:
            ct = col_types_override[col_headers[c]]
        else:
            samples = _sample_ws_column(ws, c, data_start, data_end)
            ct = _infer_col_type_from_ws(header_val, samples)
        col_types[c] = ct

        est = _estimate_col_width_from_cells(ws, c, header_row, data_end)
        # 数值/百分比列如果含长公式（>15字符），固定列宽15，避免被公式字符串撑宽
        if ct in ('money', 'number', 'pct'):
            for r in range(data_start, data_end + 1):
                txt = _cell_display_text(ws.cell(row=r, column=c))
                if txt.startswith('=') and len(txt) > 15:
                    est = 15
                    break
        ws.column_dimensions[col_letter].width = est

    # 3. 表头格式
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = s['header_font']
        cell.fill = s['header_fill']
        cell.alignment = HDR_ALIGN

    # 4. 数据行格式
    for c in range(start_col, end_col + 1):
        ct = col_types.get(c, 'unknown')
        is_num = ct in ('money', 'number', 'pct')
        align = RIGHT_ALIGN if is_num else LEFT_ALIGN
        nf = effective_fmt.get(ct)

        for r in range(data_start, data_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = align
            if nf:
                cell.number_format = nf
            # 所有单元格统一 Arial 11
            if is_num:
                txt = _cell_display_text(cell)
                # 双重检测公式：str.startswith('=') 兜不住所有情况（如 =A1+B1 存储为 '=A1+B1' 但 data_type=f）
                # 加 cell.data_type == 'f' 作为第二保险，这是 openpyxl 标记公式的官方字段
                is_formula = txt.startswith('=') or cell.data_type == 'f'
                if is_formula:
                    cell.font = s['formula_font']  # 公式=主题色黑
                else:
                    cell.font = s['data_font']     # 手动输入=主题色
            else:
                cell.font = Font(name='Arial', size=11)  # 文本=Arial 11

    # 4.5 合计行检测（提前计算，供 4.5 A列格式 和 5 段共用）
    first_val = ws.cell(row=data_end, column=1).value  # 先看 A 列
    if not first_val or not any(kw in str(first_val) for kw in SUM_KEYWORDS):
        first_val = ws.cell(row=data_end, column=start_col).value  # B 列兜底
    is_summary = (
        data_end > header_row
        and first_val is not None
        and any(kw in str(first_val) for kw in SUM_KEYWORDS)
    )

    # 4.5 A列格式补全
    # beautify 时如果 A 列有数据（如部门名称/合计行），补齐摩根系格式。
    # make_excel 模式不受影响——A列始终留空。
    a_header_val = ws.cell(row=header_row, column=1).value
    if a_header_val is not None:
        # 表头蓝底白字右对齐
        cell = ws.cell(row=header_row, column=1)
        cell.font = s['header_font']
        cell.fill = s['header_fill']
        cell.alignment = HDR_ALIGN
        # 数据行：A 列一般为文本，左对齐
        for r in range(data_start, data_end + 1):
            cell = ws.cell(row=r, column=1)
            cell.font = Font(name='Arial', size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')
        # 合计行
        if is_summary:
            cell = ws.cell(row=data_end, column=1)
            cell.font = s['sum_font']
            cell.fill = s['sum_fill']

    # 5. 合计行特殊格式（B列起的数据列）
    if is_summary:
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=data_end, column=c)
            cell.font = s['sum_font']
            cell.fill = s['sum_fill']

    # 6. 边框（规则5）：支持多表格区块自动分割
    # 扫描每行的"首个非空列"，如果连续区域的首个非空列向右偏移超过2列，
    # 则视为新的表格区块——各自应用独立框线。
    # 每个区块独立：第一行=TOP_BORDER，中间=MID_BORDER，最后=BOTTOM_BORDER
    border_start = 1 if ws.cell(row=header_row, column=1).value is not None else start_col
    
    # 计算每行的首个非空列号
    row_first_col = {}
    for r in range(header_row, data_end + 1):
        first = end_col + 1
        for c in range(border_start, end_col + 1):
            if ws.cell(row=r, column=c).value is not None:
                first = c
                break
        row_first_col[r] = first
    
    # 按首个非空列的行进偏移分割区块
    sections = []
    current_start = header_row
    current_first_col = row_first_col.get(header_row, end_col + 1)
    
    for r in range(header_row + 1, data_end + 1):
        fc = row_first_col[r]
        # 新区块条件：该行有数据 且 首个非空列比当前区块向右偏移 >2
        if fc <= end_col and fc > current_first_col + 2:
            sections.append((current_start, r - 1))
            current_start = r
            current_first_col = fc
    
    # 最后一个区块
    sections.append((current_start, data_end))
    
    # 第二步：对每个区块应用独立框线
    for sec_start, sec_end in sections:
        sec_rows = sec_end - sec_start + 1
        for c in range(border_start, end_col + 1):
            for ri, r in enumerate(range(sec_start, sec_end + 1)):
                cell = ws.cell(row=r, column=c)
                if ri == 0:
                    cell.border = TOP_BORDER
                elif ri == sec_rows - 1:
                    cell.border = BOTTOM_BORDER
                else:
                    cell.border = MID_BORDER

    # 6b. 右侧空白列
    ncols = end_col - start_col + 1
    last_data_col = start_col + ncols - 1
    right_col = get_column_letter(last_data_col + 1)
    ws.column_dimensions[right_col].width = 3

    # 7.1 条件格式自适应：将色阶最大色换为主题色
    _adapt_conditional_formatting(ws, theme)

    # 7b. 全局设置：隐藏网格线 + 冻结表头
    # freeze_rows 未指定时，根据检测到的 header_row 冻结
    ws.sheet_view.showGridLines = False
    if freeze_rows is not None:
        freeze_cell = f'A{freeze_rows + 1}' if freeze_rows > 0 else None
    else:
        freeze_cell = f'A{header_row + 1}'
    if freeze_cell:
        ws.freeze_panes = freeze_cell


def _adapt_conditional_formatting(ws, theme: str):
    """
    将条件格式中 colorScale 的最大色换为主题色。
    保持 min=白色不变，只改 max 色为表头色。
    其他类型条件格式（dataBar、iconSet 等）保持不变。

    参数
    ----
    ws : Worksheet
    theme : str — 主题名
    """
    s = _build_theme_styles(theme)
    theme_color = THEMES[theme]['header_fill']
    from openpyxl.styles import Color
    from openpyxl.formatting.rule import ColorScaleRule

    for cf in ws.conditional_formatting:
        for rule in cf.rules:
            if rule.type == 'colorScale':
                cs = rule.colorScale
                if cs.color and len(cs.color) >= 2:
                    # 替换最大色（最后一个 color）为主题色
                    cs.color[-1] = Color(rgb=theme_color)
                    cs.color[-1].type = 'rgb'


def beautify(
    input_path: str,
    output_path: Optional[str] = None,
    *,
    col_types: Optional[dict] = None,
    backup: bool = True,
    theme: str = 'default',
    fmt_override: Optional[dict] = None,
    freeze_rows: Optional[int] = None,
) -> str:
    """
    美化已有 Excel 文件，只改格式不改数据（保留公式和值）。

    参数
    ----
    input_path : str
        输入的 xlsx 文件路径。
    output_path : str, optional
        输出路径。不传则原地覆盖原文件。
    col_types : dict, optional
        手动指定列类型，如 {'订单号': 'text', '金额': 'money'}。
        覆盖自动类型推断。可选值: money/number/pct/date/text。
    backup : bool, default True
        原地覆盖前是否自动备份。备份文件名为 <原文件名>_backup_YYYYMMDD_HHMMSS.xlsx。
    theme : str
        色系主题名。可用: """ + AVAILABLE_THEMES + """。默认 'default'。
    fmt_override : dict, optional
        自定义数字格式，如 {'money': '#,##0.0', 'pct': '0.0%'}。
    freeze_rows : int, optional
        要冻结的行数。默认 None 自动推断：根据检测到的表头行冻结。
        传 0 则不冻结。

    返回
    ----
    str : 输出文件的绝对路径。

    注意
    ----
    beautify 不改任何单元格的值。公式原样保留，不转值。
    字体颜色根据公式检测自动分配：公式→黑，手动输入→蓝。
    """
    _validate_theme(theme)
    from openpyxl import load_workbook

    out = output_path or input_path

    # 自动备份——原地覆盖前自动备份原文件，防手滑
    if backup and output_path is None:
        inp = Path(input_path)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        bak = inp.with_stem(f'{inp.stem}_backup_{ts}')
        copy2(input_path, str(bak))
        print(f'备份: {bak}')

    # data_only=False（默认），公式以 '=SUM(...)' 字符串形式加载，不转值
    wb = load_workbook(input_path)

    for ws in wb.worksheets:
        _beautify_worksheet(ws, col_types_override=col_types, theme=theme, fmt_override=fmt_override,
                            freeze_rows=freeze_rows)

    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return str(out_path.absolute())


# ═══════════════════════════════════════════════════════════════════════════
# CLI 入口
#
# 两种模式：
#   python make_excel.py 数据.csv 输出.xlsx        — CSV→Excel
#   python make_excel.py --beautify 输入.xlsx 输出.xlsx  — 美化已有文件
# ═══════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='生成/美化摩根系标准格式 Excel')
    parser.add_argument('input', help='输入文件（CSV 或 --beautify 时传 xlsx）')
    parser.add_argument('output', nargs='?', default=None,
                        help='输出路径（默认: 桌面/<输入文件名>.xlsx）')
    parser.add_argument('--sheet-name', default='Sheet1', help='工作表名称（仅 CSV 模式）')
    parser.add_argument('--beautify', action='store_true',
                        help='美化已有 xlsx（只改格式不改数据，保留公式）')
    parser.add_argument('--theme', default='default', choices=list(THEMES.keys()),
                        help=f"色系主题 (default: default)。可选: {AVAILABLE_THEMES}")
    parser.add_argument('--freeze-rows', type=int, default=None,
                        help='冻结表头行数。默认自动推断。传 0 则不冻结。')
    args = parser.parse_args()

    if args.beautify:
        path = beautify(args.input, args.output, theme=args.theme, freeze_rows=args.freeze_rows)
        print(f'已美化: {path}')
        return

    if args.output is None:
        in_path = Path(args.input)
        args.output = str(Path.home() / 'Desktop' / in_path.with_suffix('.xlsx').name)

    df = pd.read_csv(args.input)
    path = make_excel(df, args.output, sheet_name=args.sheet_name, theme=args.theme,
                      freeze_rows=args.freeze_rows)
    print(f'已生成: {path}')


if __name__ == '__main__':
    main()
